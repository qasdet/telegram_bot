import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Parser:
    def __init__(self, file_path: str):
        self.__data: dict[str, dict[str, list[str]]] = {}
        self.__load_data(file_path)

    def __load_data(self, file_path: str) -> None:
        wb: Workbook = openpyxl.load_workbook(file_path)
        for index, sheet_name in enumerate(wb.sheetnames):
            # Перебор всех листов документа
            wb.active = index
            group, descipline = sheet_name.split('|')

            if descipline not in self.__data:
                self.__data[descipline] = {}

            self.__data[descipline][group] = []

            worksheet: Worksheet = wb.active
            student_name = ''
            row = 1

            while student_name is not None:
                student_name = worksheet.cell(
                    row=row, column=1).value
                if student_name is None:
                    break
                self.__data[descipline][group].append(student_name)
                row += 1

    def get_data(self) -> dict[str, dict[str, list[str]]]:
        return self.__data
